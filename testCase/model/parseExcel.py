# -- coding: utf-8 --
from openpyxl import load_workbook

#读取excel数据进行测试
class parseExcel(object):
    def __init__(self,excelPath,sheetName):
        self.wb=load_workbook(excelPath)#将要读取的excel加载到内存
        self.sheet=self.wb.get_sheet_by_name(sheetName)#通过工作表名称获取一个工作表对象
        self.maxRowNum=self.sheet.max_row#获取工作表存在数据的区域最大行号

    #返回列表数据
    def getDatasFromSheet(self):
        dataList=[]
        for line in self.sheet.rows[1:]:
            tmpList=[]
            tmpList.append(line[0].value)
            tmpList.append(line[1].value)
            dataList.append(tmpList)
        return dataList