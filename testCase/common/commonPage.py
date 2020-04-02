# -- coding: utf-8 --
from testCase.common.BasePage import BasePage
from time import sleep
from selenium.webdriver.common.by import By
import random
import xlrd
import openpyxl
from xlutils.copy import copy
import win32com.client as win32
from win32com.client import Dispatch
from testCase.model import sreenshot


class Common(BasePage):
    '''公共元素定位器'''
    # 设置定位器(class属性值中间有空格时，需要用.来代替)
    #module
    indexCen_loc = (By.ID,"tabnav_btn_0")
    newsCen_loc = (By.ID,"tabnav_btn_1")
    downLoadCen_loc = (By.ID,"tabnav_btn_2")
    filmCen_loc = (By.ID,"tabnav_btn_3")
    mallCen_loc = (By.ID,"tabnav_btn_4")
    flashCen_loc = (By.ID,"tabnav_btn_5")
    picCen_loc = (By.ID, "tabnav_btn_6")
    articleCen_loc = (By.ID,"tabnav_btn_7")
    clsInfoCen_loc = (By.ID,"tabnav_btn_8")
    #skipControl
    skipFlag_more = 1
    skipFlag_less = 1
    #text
    username_loc = (By.NAME, "username")
    password_loc = (By.NAME, "password")
    #button
    submit_loc = (By.NAME, "Submit")  # 提交
    submit2_loc = (By.NAME,"Submit2")
    exit_loc = (By.PARTIAL_LINK_TEXT,"退出")
    # vip_loc = (By.PARTIAL_LINK_TEXT, "普通会员")
    vip_loc = (By.XPATH,"//a[@href='/e/member/my/']")
    del_loc = (By.PARTIAL_LINK_TEXT, "删除")
    back_loc = (By.PARTIAL_LINK_TEXT, "返回")
    mod_loc = (By.PARTIAL_LINK_TEXT,"修改")
    #查列表的记录数
    eleText_loc = (By.XPATH, "//table[@class='tableborder'][2]/tbody/tr")
    # 全局变量
    randoma = 0     #整数随机数
    tempVar = ""    #字符串型
    #断言标志（全局变量），主要用于判断是否增加或者删除成功
    assertTip=""
    #列表查询（针对商品）
    new_url = ""
    open_url=""
    # 图片上传路径
    userpicfile_path = r'F:\Python\phomeNetFront\test_data\pic\0.jpg'  # 默认上传图片
    softwareUpload_path = r'F:\Python\phomeNetFront\test_data\software\friend.rar'
    flashUpload_path = r'F:\Python\phomeNetFront\test_data\flash\flash.swf'
    #Selector
    cid_loc = (By.NAME, "cid")
    cidOp_loc = (By.XPATH, "//select[@name='cid']/option")

    # 注册时，获取一个随机数作为注册账号；或者其他功能获取随机数所用
    def type_getAccount(self):
        self.randoma = random.randint(10000000, 99999999)
        return self.randoma

    # 定义username输入框元素的定位、输入方法
    def type_username(self,username):
        self.find_element(*self.username_loc).clear() # 调用find_element()方法，username定位器变量作为参数传入
        self.find_element(*self.username_loc).send_keys(username) # 输入username参数值

    # 定义password输入框元素的定位、输入方法
    def type_password(self,password):
        self.find_element(*self.password_loc).clear()
        self.find_element(*self.password_loc).send_keys(password)

    '''登录类'''
    # 定义Submit按钮元素的点击方法
    def type_submit(self):
        self.find_element(*self.submit_loc).click()

    # 封装登录功能模块的方法
    def login_action(self,username,password):
        try:
            #
            print("\033[1;31;40m 登录页面正在打开...\033[0m")
            self.u_para = ""
            self.open()
            self.type_username(username)
            self.type_password(password)
            self.type_submit()
        except BaseException as msg:
            print(msg)

    #系统退出登录
    def type_sysExit(self):
        try:
            self.find_element(*self.exit_loc).click()
            self.type_getAlert()
            self.driver.refresh()
        except BaseException as msg:
            print(msg)

    '''get()方法定义区'''
    def getTotalPage(self):
        total_pages=len(self.driver.find_element_by_class_name("pagination").find_elements_by_tag_name("li")) - 4
        print("\033[1;34;0m total_pages is：%d\033[0m"%total_pages)
        return total_pages

    #一般用link_text寻找元素，看是否找到并返回text
    def type_getText(self,target_loc):
        try:
            sleep(3)
            text = self.find_element(*target_loc).text
            return text
        except BaseException as msg:
            print(msg)

    def type_getAlert(self):
        try:
            alert = self.driver.switch_to.alert
            sleep(1)
            alert.accept()
        except BaseException as msg:
            print(msg)

    '''excel操作方法'''
    #读取excel列表中第一列的数据
    def get_excelList(self,excelName):
        try:
            #打开Excel
            workbook = xlrd.open_workbook('F:\\Python\\phomeNetFront\\test_data\\excel\\'+excelName)
            print("\033[1;33;0m %s 打开已执行\033[0m"%excelName)
            # 进入sheet
            excel_sheet = workbook.sheet_by_index(0)
            # 获取行数和列
            max_row = excel_sheet.nrows
            data=[]
            row_num = 0
            while row_num < max_row:
                 # 将表中第一列的所有数据写入data数组中
                data.append(excel_sheet.cell(rowx=row_num,colx=0).value)
                row_num = row_num + 1
            print("data is: %r"%data)
            workbook.release_resources()

            # workbook = openpyxl.load_workbook('F:\\Python\\phomeNetFront\\test_data\\excel\\' + excelName)
            # worksheet = workbook.worksheets[0]
            # max_row = worksheet.nrows
            # print("getExcelRow's max_row is:%d" % max_row)
            # data = []
            # row_num = 1
            # while row_num <= max_row:
            #     # 将表中第一列的所有数据写入data数组中
            #     data.append(worksheet.cell(rowx=row_num, colx=1).value)
            #     row_num = row_num + 1
            # print("data is: %r" % data)
            # workbook.save('F:\\Python\\phomeNetFront\\test_data\\excel\\' + excelName)
            return data
        except BaseException as msg:
            print(msg)

    #获取列表中第一列的行数
    def get_ExcelRow(self,excelName):
        try:
            # 打开Excel
            workbook = xlrd.open_workbook('F:\\Python\\phomeNetFront\\test_data\\excel\\' + excelName)
            print("\033[1;33;0m %s 获取行数正在执行\033[0m" % excelName)
            # 进入sheet
            excel_sheet = workbook.sheet_by_index(0)
            # 获取行数
            max_row = excel_sheet.nrows
            print("getExcelRow's max_row is:%d"%max_row)

            # workbook = openpyxl.load_workbook('F:\\Python\\phomeNetFront\\test_data\\excel\\' + excelName)
            # worksheet = workbook.worksheets[0]
            # max_row = worksheet.nrows
            # print("getExcelRow's max_row is:%d" % max_row)
            # workbook.save('F:\\Python\\phomeNetFront\\test_data\\excel\\' + excelName)

            return max_row
        except BaseException as msg:
            print(msg)

    def get_ranExcelText(self,excelName):
        try:
            # 打开Excel
            workbook = xlrd.open_workbook('F:\\Python\\phomeNetFront\\test_data\\excel\\' + excelName)
            print("\033[1;33;0m %s 获取行数随机数正在执行\033[0m" % excelName)
            # 进入sheet
            excel_sheet = workbook.sheet_by_index(0)
            # 获取行数
            max_row = excel_sheet.nrows
            print("get_ranExcelText maxrow is:%d"%max_row)
            ranRow=random.randint(1,max_row)-1
            print("ranRow is:%d"%ranRow)
            text=excel_sheet.cell(rowx=ranRow, colx=0).value
            print("get_ranExcelText is:%r"%text)
            workbook.release_resources()

            # workbook = openpyxl.load_workbook('F:\\Python\\phomeNetFront\\test_data\\excel\\' + excelName)
            # worksheet = workbook.worksheets[0]
            # print("\033[1;33;0m %s 获取行数随机数正在执行\033[0m" % excelName)
            # max_row = worksheet.nrows
            # print("get_ranExcelText maxrow is:%d" % max_row)
            # ranRow=random.randint(1,max_row)
            # print("ranRow is:%d"%ranRow)
            # text=worksheet.cell(rowx=ranRow, colx=1).value
            # print("get_ranExcelText is:%r"%text)
            # workbook.save('F:\\Python\\phomeNetFront\\test_data\\excel\\' + excelName)
            return (ranRow,text)
        except BaseException as msg:
            print(msg)

    def get_ExcelText(self,excelName,trow):
        try:
            # 打开Excel
            workbook = xlrd.open_workbook('F:\\Python\\phomeNetFront\\test_data\\excel\\' + excelName)
            print("\033[1;33;0m %s 根据row获取文本值正在执行\033[0m" % excelName)
            # 进入sheet
            excel_sheet = workbook.sheet_by_index(0)
            # 获取行数
            text=excel_sheet.cell(rowx=trow, colx=0).value
            print("get_ranExcelText is:%r"%text)
            workbook.release_resources()

            # workbook = openpyxl.load_workbook('F:\\Python\\phomeNetFront\\test_data\\excel\\' + excelName)
            # worksheet = workbook.index(0)
            # print("\033[1;33;0m %s 根据row获取文本值正在执行\033[0m" % excelName)
            # text = worksheet.cell(rowx=trow, colx=1).value
            # print("get_ranExcelText is:%r" % text)
            # workbook.save('F:\\Python\\phomeNetFront\\test_data\\excel\\' + excelName)
            return text
        except BaseException as msg:
            print(msg)

    #获取列表记录数-1
    def get_listCount(self,target_loc):
        try:
            sleep(2)
            length = len(self.find_elements(*target_loc))
            print("type_count length is：%d" % length)
            return (length - 1)
        except BaseException as msg:
            print(msg)

    #截图
    def get_sreenshot(self):
        try:
            return sreenshot.sreenshot
        except BaseException as msg:
            print(msg)

    #切换到最后一个handle
    def type_getLastHandle(self):
        try:
            sleep(2)
            handles = self.driver.window_handles
            print("index_handle is:%r" % self.index_handle)
            print("handles are :%r" % handles)
            self.driver.switch_to.window(handles[-1])
            return handles[-1]
        except BaseException as msg:
            print(msg)

    #关闭窗口，切回到第一个handle
    def type_closeHandle(self):
        try:
            self.driver.close()
            self.driver.switch_to.window(self.index_handle)
        except BaseException as msg:
            print(msg)

    '''set()方法定义区'''
    #将数据追加到excel表中
    def append_excel(self,excelName,var):
        try:
            #打开Excel
            workbook = xlrd.open_workbook('F:\\Python\\phomeNetFront\\test_data\\excel\\'+excelName)
            print("\033[1;33;0m %s 追加正在执行\033[0m"%excelName)

            # 进入sheet
            excel_sheet = workbook.sheet_by_index(0)
            max_row = excel_sheet.nrows

            new_workbook = copy(workbook)  # 将xlrd对象拷贝转化为xlwt对象
            new_worksheet = new_workbook.get_sheet(0)  # 获取转化后工作簿中的第一个表格
            new_worksheet.write(max_row, 0, var)  # 追加写入数据
            print("追加的数据为：%s"%var)
            new_workbook.save('F:\\Python\\phomeNetFront\\test_data\\excel\\'+excelName)  # 保存工作簿
            workbook.release_resources()
            print("xls格式表格追加数据成功！")
            self.get_excelList(excelName)
        except BaseException as msg:
            print(msg)

    def mod_excel(self, excelName, var):
        try:
            # 打开Excel
            workbook = xlrd.open_workbook('F:\\Python\\phomeNetFront\\test_data\\excel\\' + excelName)
            print("\033[1;33;0m %s 修改正在执行\033[0m" % excelName)

            # 进入sheet
            excel_sheet = workbook.sheet_by_index(0)
            max_row = excel_sheet.nrows

            new_workbook = copy(workbook)  # 将xlrd对象拷贝转化为xlwt对象
            new_worksheet = new_workbook.get_sheet(0)  # 获取转化后工作簿中的第一个表格
            new_worksheet.write(max_row-1, 0, var)  # 修改数据
            new_workbook.save('F:\\Python\\phomeNetFront\\test_data\\excel\\' + excelName)  # 保存工作簿
            workbook.release_resources()
            self.get_excelList(excelName)
        except BaseException as msg:
            print(msg)

    #删除列表中匹配的数据，修改删除excel的方法
    def del_excel(self,excelName,row):
        try:
            # 打开Excel
            workApp = win32.Dispatch('Excel.Application')
            workbook=workApp.Workbooks.Open('F:\\Python\\phomeNetFront\\test_data\\excel\\' + excelName)
            sht = workbook.Worksheets("Sheet1")
            sht.Rows(row).Delete()
            workbook.Save()
            workbook.Close(SaveChanges=0)
            self.append_excel(excelName,"")
            self.get_excelList(excelName)
            # workbook = openpyxl.load_workbook('F:\\Python\\phomeNetFront\\test_data\\excel\\' + excelName)
            # worksheet = workbook.worksheets
            # print("\033[1;33;0m %s 删除正在执行\033[0m" % excelName)
            # worksheet.remove()
            # workbook.save('F:\\Python\\phomeNetFront\\test_data\\excel\\' + excelName)

        except BaseException as msg:
            print(msg)

    #打开普通会员
    def type_openVip(self):
        try:
            print("\033[1;31;40m 打开普通会员管理页面 \033[0m")
            self.find_element(*self.vip_loc).click()
            sleep(2)
            handles = self.driver.window_handles
            self.driver.switch_to.window(handles[-1])
        except BaseException as msg:
            print(msg)

